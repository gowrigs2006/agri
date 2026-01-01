from flask import Flask, request, jsonify
from werkzeug.utils import secure_filename
from flask_cors import CORS
from routes.crop import crop_bp
from routes.weather import weather_bp
import tensorflow as tf
import numpy as np
import os
import uuid
import json
from openpyxl import Workbook, load_workbook
from tensorflow.keras.applications.efficientnet import preprocess_input

# Initialize Flask App
app = Flask(__name__)
CORS(app)

# Register Blueprints
app.register_blueprint(crop_bp)
app.register_blueprint(weather_bp)

# Load ML Model and Metadata
pest_model = tf.keras.models.load_model("models/pest_model.keras")

with open("models/class_indices.json", "r") as f:
    class_indices = json.load(f)

with open("models/pest_advice.json", "r") as f:
    pest_advice = json.load(f)

# 🌾 Submit Crop Sale
@app.route("/submit-sale", methods=["POST"])
def submit_sale():
    data = request.get_json()
    crop = data.get("crop")
    amount = data.get("amount")
    phone = data.get("phone")

    filename = "sales_data.xlsx"

    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append(["Crop", "Amount (kg)", "Phone"])
        wb.save(filename)

    wb = load_workbook(filename)
    ws = wb.active
    ws.append([crop, amount, phone])
    wb.save(filename)

    return f"✅ Sale recorded: {amount} kg of {crop} from {phone}"

# 🪲 Pest Detection
@app.route("/detect", methods=["POST"])
def detect_pest():
    if "image" not in request.files:
        return jsonify({"error": "No image uploaded"}), 400

    image = request.files["image"]
    if image.filename == "":
        return jsonify({"error": "Empty filename"}), 400

    os.makedirs("uploadimages", exist_ok=True)
    safe_filename = secure_filename(image.filename)
    filepath = f"uploadimages/temp_{uuid.uuid4().hex}_{safe_filename}"
    image.save(filepath)

    try:
        print("🔍 Predicting pest from:", filepath)

        # Load and preprocess using EfficientNetB4 preprocessing
        img = tf.keras.utils.load_img(filepath, target_size=(160, 160))
        arr = tf.keras.utils.img_to_array(img)
        arr = np.expand_dims(arr, axis=0)
        arr = preprocess_input(arr)  # ✅ proper preprocessing for EfficientNetB4

        preds = pest_model.predict(arr)
        class_index = np.argmax(preds)

        class_name = list(class_indices.keys())[list(class_indices.values()).index(class_index)]
        print("🔢 Predicted class:", class_name)

        match = next((item for item in pest_advice if item["name"] == class_name), None)
        advice = match["cure"] if match else "No advice available."

        return jsonify({
            "pestName": class_name,
            "solution": advice
        })

    except Exception as e:
        print("🔥 ERROR during prediction:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if os.path.exists(filepath):
            os.remove(filepath)


# ✅ Index
@app.route("/")
def index():
    return "🌿 AgriGuru Unified Backend Running"

if __name__ == "__main__":
    app.run(debug=True)
